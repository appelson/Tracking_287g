# Importing packages
import os
import glob
from openpyxl import load_workbook

# defining a funciton
def excel_to_readme(dataset):
    if dataset == "participating":
        path_pattern = os.path.expanduser('results/results_*/participatingAgencies*.xlsx')
    elif dataset == "pending":
        path_pattern = os.path.expanduser('results/results_*/pendingAgencies*.xlsx')
    else:
        return "Invalid dataset type"
    
    excel_files = glob.glob(path_pattern)
    excel_files.sort(key=lambda x: x.split('participatingAgencies')[1].split('.')[0], reverse=True)

    recent_excel_file = excel_files[0]
    wb = load_workbook(recent_excel_file)
    sheet = wb.active

    columns_to_keep = ["STATE", "LAW ENFORCEMENT AGENCY", "TYPE", "COUNTY", "SUPPORT TYPE", "SIGNED"]

    header = [cell.value for cell in sheet[1]]

    column_indices = [header.index(col) for col in columns_to_keep]

    table = []
    for row in sheet.iter_rows(values_only=True):
        filtered_row = [row[i] for i in column_indices]
        table.append(filtered_row)

    markdown = "| " + " | ".join(str(cell) for cell in table[0]) + " |\n"
    markdown += "| " + " | ".join("---" for _ in table[0]) + " |\n"

    for row in table[1:]:
        markdown += "| " + " | ".join(str(cell) for cell in row) + " |\n"

    return markdown

markdown_table = excel_to_readme("participating")

readme_path = os.path.expanduser('sheets/README.md')

with open(readme_path, 'w') as f:
    f.write(markdown_table)
